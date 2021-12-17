import openpyxl

def getRowCount(filePath,sheetName):
    workbook=openpyxl.load_workbook(filePath)
    sheet=workbook[sheetName]
    rows=sheet.max_row
    return rows

def getColumnCount(filePath,sheetName):
    workbook=openpyxl.load_workbook(filePath)
    sheet=workbook[sheetName]
    columns=sheet.max_columns
    return columns

def readExcel(filePath,sheetName,rownum,colnum):
    workbook=openpyxl.load_workbook(filePath)
    sheet=workbook[sheetName]
    data=sheet.cell(row=rownum, column=colnum).value
    return data


def writeExcel(filePath,sheetName,rownum,colnum,data):
    workbook=openpyxl.load_workbook(filePath)
    sheet=workbook[sheetName]
    sheet.cell(row=rownum,column=colnum).value=data
    workbook.save(filePath)

