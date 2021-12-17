# To read data from excel file

import openpyxl
import os
import time

cwd=os.getcwd()
path=cwd+"/"+"ProjectFiles"+"/"+"ReadFile.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook["Sheet1"]

rows=sheet.max_row
cols=sheet.max_column

print("Number of rows=",rows)
print("Number of columns=",cols)

# print data from excel

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="         ")
    print()
